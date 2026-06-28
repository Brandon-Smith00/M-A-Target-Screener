# ─────────────────────────────────────────────
#  main.py — M&A Target Screener Orchestrator
# ─────────────────────────────────────────────

import sys
import os
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

from config import SECTOR_UNIVERSES, TOP_N_TARGETS, OUTPUT_DIR, SUMMARY_FILENAME
from data_fetch import fetch_sector_universe, format_billions, format_pct, format_multiple
from scorer import score_all, get_score_label
from ai_analyst import enrich_with_commentary
from tearsheet import generate_tearsheet

NAVY_HEX  = "0A2342"
LGRAY_HEX = "F5F6F7"
GREEN_HEX = "177345"
AMBER_HEX = "D4A017"
RED_HEX   = "B22222"

def score_fill(score):
    if score >= 80: h = GREEN_HEX
    elif score >= 65: h = "20A160"
    elif score >= 50: h = AMBER_HEX
    elif score >= 35: h = "D96813"
    else: h = RED_HEX
    return PatternFill("solid", fgColor=h)

def build_excel_summary(df, sector):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, SUMMARY_FILENAME)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Targets"

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"M&A Target Screener — {sector.replace('_',' ').title()} | {date.today().strftime('%B %d, %Y')}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY_HEX)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Rank","Ticker","Company","Industry","Mkt Cap","EV","Revenue","EBITDA","EBITDA Mg.","Rev. Growth","EV/EBITDA","Net D/EBITDA","Total Score","Rating"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    for i, (_, row) in enumerate(df.head(TOP_N_TARGETS * 2).iterrows()):
        r = i + 3
        bg = LGRAY_HEX if i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)
        data = [row.get("rank",i+1), row.get("ticker",""), row.get("name","")[:40], row.get("industry","")[:30],
                format_billions(row.get("market_cap")), format_billions(row.get("enterprise_value")),
                format_billions(row.get("revenue")), format_billions(row.get("ebitda")),
                format_pct(row.get("ebitda_margin")), format_pct(row.get("revenue_growth")),
                format_multiple(row.get("ev_ebitda")), format_multiple(row.get("nd_ebitda")),
                row.get("total_score", 0), get_score_label(row.get("total_score", 0))]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
        score = row.get("total_score", 0)
        for sc in [13, 14]:
            ws.cell(row=r, column=sc).fill = score_fill(score)
            ws.cell(row=r, column=sc).font = Font(size=9, bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 16

    for col, w in enumerate([6,8,32,28,10,10,10,10,11,11,10,12,12,12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"
    wb.save(path)
    return path


def run_screener(sector="digital_infrastructure", gemini_key=""):
    print("\n" + "="*60)
    print("  AI-POWERED M&A TARGET SCREENER")
    print("  Built with Gemini + yfinance")
    print("="*60)

    start = time.time()
    df = fetch_sector_universe(sector)
    if df.empty:
        print("❌  No data fetched.")
        return

    df = score_all(df)

    if gemini_key:
        df = enrich_with_commentary(df, api_key=gemini_key, top_n=TOP_N_TARGETS)
    else:
        print("⚠  No API key found in config.py — skipping AI commentary.\n")
        df["commentary"] = None

    print(f"📄  Generating PDF tearsheets for top {TOP_N_TARGETS} targets...")
    for _, row in df.head(TOP_N_TARGETS).iterrows():
        path = generate_tearsheet(row.to_dict(), rank=int(row["rank"]))
        print(f"     ✓ {os.path.basename(path)}")

    print(f"\n📊  Building Excel summary...")
    excel_path = build_excel_summary(df, sector)
    print(f"     ✓ {os.path.basename(excel_path)}")

    elapsed = time.time() - start
    print(f"\n✅  Done in {elapsed:.1f}s  |  Outputs saved to ./{OUTPUT_DIR}/")
    print("="*60 + "\n")
    return df


if __name__ == "__main__":
    sector = sys.argv[1] if len(sys.argv) > 1 else "digital_infrastructure"
    if sector not in SECTOR_UNIVERSES:
        print(f"❌  Unknown sector. Available: {', '.join(SECTOR_UNIVERSES.keys())}")
        sys.exit(1)

    # Read Gemini key from config
    try:
        from config import GEMINI_API_KEY
    except ImportError:
        GEMINI_API_KEY = ""

    run_screener(sector, gemini_key=GEMINI_API_KEY)
